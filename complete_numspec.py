import pandas as pd
import openpyxl
import warnings
import re
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# Load the workbook and select the sheet
wb = openpyxl.load_workbook('QBOT21000.xlsx', data_only=True)
ws = wb['Bom']

# Find the PivotTable named 'BomOracle'
pivot = None
for pt in ws._pivots:
    if pt.name == 'BomOracle':
        pivot = pt
        break

if pivot is None:
    raise ValueError("PivotTable 'BomOracle' not found.")

# Get the pivot table's displayed range (where it is rendered in the sheet)
pivot_display_range = pivot.location.ref

# Extract the displayed data from the range
data = []
for row in ws[pivot_display_range]:
    data.append([cell.value for cell in row])

# Convert to pandas DataFrame
df = pd.DataFrame(data[1:], columns=data[0])
# Filter rows where 'Item' matches the pattern '344********_**'
pattern = re.compile(r'^344\d{5}_\d{2}$')
df = df[df['Item'].apply(lambda x: bool(pattern.match(str(x))))].reset_index(drop=True)


# Load the 'Numero Spec.xlsm' workbook and the sheet containing 'Tableau1'
file_path = 'C:/OneDrive - Schneider Electric/Documents - MBT Plant - Gestion affaires - Métiers Group/Metiers/BE/Data Bases/N° Specifications/Numero Spec.xlsm'
wb_num = openpyxl.load_workbook(file_path, keep_vba=True)
ws_num = wb_num['Numéros de Spec']
table = None

# Find the table named 'Tableau1' in the specified sheet
table = ws_num.tables['Tableau1']

if table is None:
    raise ValueError("Table 'Tableau1' not found in 'Numéro de Spec' sheet of 'Numero Spec.xlsm'.")

# Get the table's range and data
table_range = table.ref
table_data = []
for row in ws_num[table_range]:
    table_data.append([cell.value for cell in row])

# Convert to DataFrame for easier manipulation
df_num = pd.DataFrame(table_data[1:], columns=table_data[0]).dropna(subset=['NUMERO']).reset_index(drop=True)

item_prefixes = df['Item'].astype(str).str.slice(0, -3).astype(int).tolist()
# print(item_prefixes)
# print(df_num['NUMERO'].dropna().astype(int).head(10))

mask = df_num['NUMERO'].astype(int).isin(item_prefixes)

# df_num.loc[mask, 'DESIGNATION'] = row['Description']

for idx, row in df_num.loc[mask].iterrows():
    match = df[df['Item'].astype(str).str[:-3] == str(int(row['NUMERO']))]
    # print(str(int(row['NUMERO'])))
    # print(match)
    if not match.empty:
        df_num.at[idx, 'DESIGNATION'] = match.iloc[0]['Description']

# print(df['Description'].head(10))

# print(df_num[['NUMERO', 'DESIGNATION']].head(10))

# Find the starting row of the table 'Tableau1' in the worksheet +1 to skip header row
table_start_row = ws_num[table.ref.split(':')[0]].row + 1


# Write back the updated DESIGNATION column to the worksheet,
# aligning DataFrame rows to Excel table rows
designation_col_idx = df_num.columns.get_loc('DESIGNATION') + ws_num[table.ref.split(':')[0]].column

for df_idx, row in df_num.loc[mask].iterrows():
    excel_row = table_start_row + df_idx
    ws_num.cell(row=excel_row, column=designation_col_idx, value=row['DESIGNATION'])
    print(f'Changed description of {int(row['NUMERO'])} to {row['DESIGNATION']}')

# Save the workbook
wb_num.save(file_path)
