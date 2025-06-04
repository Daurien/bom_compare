import argparse
import math
import os
from pprint import pprint
import re
import warnings
# import pandas as pd
from pandas import DataFrame
from pandas.io.excel import read_excel as pd_read_excel
from deepdiff import DeepDiff
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.packaging.custom import BoolProperty, DateTimeProperty, StringProperty, CustomPropertyList, IntProperty
import shutil
import os
import pandas as pd
from openpyxl import load_workbook
import warnings
from numpy import ndarray
import tkinter as tk
from tkinter import font


def read_bom_to_numpy(filePath: str) -> ndarray:
    """
    Reads an Excel file (ORACLE or CREO type) and returns specified BOM columns as a NumPy array.

    Args:
        filePath (str): Path to the Excel file

    Returns:
        np.ndarray: Array containing BOM data with selected columns
    """
    columns_to_find = ['Level', 'Item', 'Description', 'Quantity', 'Supply Type', 'Supplier_Type', 'SE_REVISION']
    dtypes = {'Level': int, 'Item': str, 'Description': str, 'Quantity': int, 'Supply Type': str}

    # Determine file type
    workbook = load_workbook(filePath, read_only=True)
    file_type = 'CREO' if 'Bom' in workbook.sheetnames and 'Import_Creo' in workbook.sheetnames else 'ORACLE'
    workbook.close()

    # Find table origin
    workbook = load_workbook(filePath)
    sheet_name = 'Import_Creo' if file_type == 'CREO' else None
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    skip = None
    for table in worksheet.tables.values():
        origin_cell = table.ref.split(':')[0]
        skip = int(''.join(filter(str.isdigit, origin_cell))) - 1
        break

    workbook.close()

    if skip is None:
        raise ValueError("No table found in the Excel file")

    # Read and process BOM
    with warnings.catch_warnings(action="ignore"):
        if file_type == 'CREO':
            BOM = pd.read_excel(filePath, sheet_name='Import_Creo', dtype=dtypes, skiprows=skip)
        else:
            BOM = pd.read_excel(filePath, dtype=dtypes, skiprows=skip)

        columns_to_keep = [col for col in columns_to_find if col in BOM.columns]
        return BOM[columns_to_keep].to_numpy()


def bom_excel_to_dictionary(filePath: str):
    """
    Convert a BOM Excel file to a dictionary.

    Args:
        path (str): The file path to the BOM Excel file.

    Returns:
        tuple: A tuple containing the BOM dictionary and metadata.
    """
    # read the excel BOM file ignoring the first row and convert it to a numpy array
    # with warnings.catch_warnings(action="ignore"):
    #     file_type = get_file_type(filePath)
    #     columns_to_find = ['Level', 'Item', 'Description', 'Quantity', 'Supply Type', 'Supplier_Type', 'SE_REVISION']
    #     if file_type == 'ORACLE':
    #         skip = find_table_origin_line_number(filePath)-1
    #         BOM = pd_read_excel(filePath, dtype={'Level': int, 'Item': str, 'Description': str,
    #                             'Quantity': int, 'Supply Type': str}, skiprows=skip)
    #         columns_to_keep = [element for element in columns_to_find if element in BOM.keys()]
    #         BOM = BOM[columns_to_keep].to_numpy()
    #     elif file_type == 'CREO':
    #         skip = find_table_origin_line_number(filePath, 'Import_Creo')-1
    #         BOM = pd_read_excel(filePath, sheet_name='Import_Creo', dtype={'Level': int, 'Item': str, 'Description': str,
    #                             'Quantity': int, 'Supply Type': str}, skiprows=skip)
    #         columns_to_keep = [element for element in columns_to_find if element in BOM.keys()]
    #         BOM = BOM[columns_to_keep].to_numpy()
    BOM = read_bom_to_numpy(filePath)

    # # unnify dat format to string
    # for row in BOM:
    #     if isinstance(row[10], datetime):
    #         row[10] = row[10].strftime('%m/%d/%Y')
    max_depth = max(BOM[:, 0])
    bom_length = len(BOM)
    last_row = bom_length - 1

    # # delete all useless columns
    # BOM = np.delete(BOM, [max_depth+6], 1)
    # BOM = np.delete(BOM, np.s_[max_depth+7:], 1)

    contents = []
    prev_contents = []
    current_content = []
    reading = False
    final_bom = {}

    for col in range(max_depth, 0, -1):
        # for deepest level
        if col == max_depth and col != 1:
            for row in range(bom_length):
                if BOM[row, 0] == col:
                    reading = True
                    current_content.append(append_row(BOM[row]))

                if row == last_row or reading:
                    reading = False
                    contents.append(current_content)
                    current_content = []

            prev_contents = contents
            contents = []

        # for intermediate levels
        elif col > 1:
            for row in range(bom_length):
                if BOM[row, 0] == col:
                    reading = True
                    try:
                        if BOM[row+1, 0] == col+1:  # if current row contains sub-level items
                            current_content.append(append_row(BOM[row], prev_contents.pop(0)))
                        else:  # if current does not contains sub-level items
                            current_content.append(append_row(BOM[row]))
                    except IndexError:  # if we are reading last row
                        current_content.append(append_row(BOM[row]))

                if row == last_row or (reading and BOM[row, 0] < col):
                    reading = False
                    contents.append(current_content)
                    current_content = []

            prev_contents = contents
            contents = []

        # for top level
        else:
            for row in range(bom_length):
                if BOM[row, 0] == col and row != last_row:
                    try:
                        if BOM[row+1, 0] == col+1:  # if current row contains sub-level items
                            current_content.append(append_row(BOM[row], prev_contents.pop(0)))
                        else:  # if current does not contains sub-level items
                            current_content.append(append_row(BOM[row]))
                    except IndexError:  # if we are reading last row
                        current_content.append(append_row(BOM[row]))

                elif row == last_row:
                    final_bom = array_to_dict(current_content)
                    # final_bom = {'BOM''Item name': 'QSMA', 'content': array_to_dict(current_content)}

    return final_bom, max_depth


def light_bom_excel_to_dictionary(filePath: str):
    """Take BOM from 'BOM' sheet of Creo extraceted Excel file and return a simplified BOM dictionary.

    Args:
        filePath (str): file to read (from Creo only)

    Raises:
        ValueError: If pivot table 'BOmOracle' is not found in the workbook.

    Returns:
        dict: bom transformed as a dictionary with 'Item name' as keys and other attributes as values.
    """
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    # Load the workbook and select the sheet
    wb = load_workbook(filePath, data_only=True)
    try:
        ws = wb['Bom']
    except KeyError:
        raise ValueError("Sheet 'BOM' not found in the workbook.")

    # Find the PivotTable named 'BomOracle'
    pivot = None
    for pt in ws._pivots:
        if pt.name == 'BomOracle':
            pivot = pt
            break

    if pivot is None:
        raise ValueError("PivotTable 'BomOracle' not found.")

    # Get the pivot table's displayed range (where it is rendered in the sheet)
    pivot_display_range = pivot.location.ref

    # Extract the displayed data from the range
    data = []
    for row in ws[pivot_display_range]:
        data.append([cell.value for cell in row])

    # Convert to pandas DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])
    df = df[df['Item'].notna()].reset_index(drop=True)

    # Remove suffix pattern from 'Item' column
    df['Item'] = df['Item'].str.replace(r'_\d{2}$', '', regex=True)

    # Rename columns
    df = df.rename(columns={'Supplier_Type': 'SupplyType', 'Item': 'Item name',
                   'SE_REVISION': 'Revision', 'Qty': 'Quantity'})

    # Keep only columns that are different from the original and not unchanged
    unchanged_cols = set(df.columns) - set(['SupplyType', 'Item name', 'Revision', 'Quantity', 'Description'])
    df = df[['SupplyType', 'Item name', 'Revision', 'Quantity', 'Description']]

    # Add 'Depth' column full of 1
    df['Depth'] = 1

    df['SupplyType'] = ""

    df['Revision'] = df['Revision'].apply(lambda x: f"{int(x):02d}" if x and str(x).isdigit() else x)

    light_bom = {}

    for _, row in df.iterrows():
        item = row['Item name']
        light_bom[item] = row.to_dict()

    print(f"Quantity column type: {df['Quantity'].dtype}")

    return light_bom


def light_bom_oracle_to_dictionary(filePath: str):
    """Take txt file BOM extraceted from Oracle and return a simplified BOM dictionary.

    Args:
        filePath (str): file to read (from Oracle only)

    Returns:
        dict: bom transformed as a dictionary with 'Item name' as keys and other attributes as values.
    """
    # Read the tab-separated file into a DataFrame
    df = pd.read_csv(filePath, sep='\t')

    # Create the simplified DataFrame
    simplified_df = df[['Level', 'Supply Type', 'Item', 'Revision', 'Quantity', 'Description']].rename(
        columns={
            'Supply Type': 'SupplyType',
            'Item': 'Item name',
            'Revision': 'Revision',
            'Quantity': 'Quantity',
            'Description': 'Description'
        })
    # Clean whitespace from all string columns
    simplified_df = simplified_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # Convert Quantity column to float64
    simplified_df['Quantity'] = simplified_df['Quantity'].astype('float64')

    # Add 'Depth' column full of 1
    simplified_df['Depth'] = 1
    simplified_df['Revision'] = "XX"  # Default revision value
    simplified_df['SupplyType'] = ""

    # Extract revision from 'Item name' column and update Revision column

    # Step 1: Create a buffer column with tuples (cleaned_name, revision)
    buffer = simplified_df['Item name'].apply(lambda x: (
        x[:-3], x[-2:]) if re.search(r'_\d{2}$', str(x)) else (x, simplified_df.loc[simplified_df['Item name'] == x, 'Revision'].iloc[0]))

    # Step 2: Update the 'Revision' column from the buffer
    simplified_df['Revision'] = buffer.apply(lambda x: f"{int(x[1]):02d}" if x[1] and str(x[1]).isdigit() else x[1])

    # Step 3: Update the 'Item name' column from the buffer
    simplified_df['Item name'] = buffer.apply(lambda x: x[0])

    # Find parent for each row
    parent_list = []
    for idx, row in simplified_df.iterrows():
        current_level = row['Level']
        if current_level == 1:
            parent = "main"
        else:
            parent = None
            # Look upwards for the most recent row with Level == current_level - 1
            for prev_idx in range(idx - 1, -1, -1):
                if simplified_df.loc[prev_idx, 'Level'] == current_level - 1:
                    parent = simplified_df.loc[prev_idx, 'Item name']
                    break
        parent_list.append(parent)

    simplified_df['parent'] = parent_list

    # Add category column based on the logic
    category_list = []
    for idx, row in simplified_df.iterrows():
        current_item = str(row['Item name'])
        parent_item = str(row['parent'])
        # Check next row's Item name if it exists
        if idx + 1 < len(simplified_df):
            next_item_parent = str(simplified_df.loc[idx + 1, 'parent'])
        else:
            next_item_parent = None
        if (next_item_parent == current_item) and (current_item != parent_item):
            category = "Assembly"
        else:
            category = "Part"
        category_list.append(category)

    simplified_df['category'] = category_list

    # Filter out rows where category is 'Assembly'
    simplified_df = simplified_df[simplified_df['category'] != 'Assembly'].drop(
        ['parent', 'category', 'Level'], axis=1)

    light_bom = {}

    for _, row in simplified_df.iterrows():
        item = row['Item name']
        light_bom[item] = row.to_dict()

    return light_bom


def get_file_type(filePath: str):
    workbook = load_workbook(filePath, read_only=True)
    sheet_names = workbook.sheetnames

    if 'Bom' in sheet_names and 'Import_Creo' in sheet_names:
        return 'CREO'
    else:
        return 'ORACLE'


def array_to_dict(array):
    res = {}

    for row in array:
        res[row['Item name']] = row

    return res


def find_table_origin_line_number(file_path, sheet_name=None):
    # Load the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    if sheet_name:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    # Iterate through the tables to find the origin
    for table in worksheet.tables.values():
        origin_cell = table.ref.split(':')[0]
        # Extract the row number from the cell reference
        row_number = int(''.join(filter(str.isdigit, origin_cell)))

        return row_number

    return None


def append_row(row, content=None):

    # search pattern like '*_02' in name to extract revision
    item_name = str(row[1])
    pattern = r'_(\d{2})$'
    match = re.search(pattern, item_name)

    if match:
        revision = match.group(1)
        item_name = item_name[:-3]
    elif len(row) == 6:
        revision = row[5]
    else:
        revision = 'xx'

    result = {
        'Item name': item_name,
        'Revision': revision,
        'Description': row[2],
        'Quantity': row[3],
        # 'FromDate': row[max_depth+5],
        'SupplyType': row[4],
        'Depth': row[0]
    }

    if content:
        result['content'] = array_to_dict(content)

    return result


def append_to_dict(keys: list, bom_content: dict, modify_type: str, initial_dict: dict = None):
    """_summary_

    Args:
        keys (list): list of key string EX :  ['34410718','34410662','34411697'] that represent the 'path' to the final level of '34411697' (subcomponent of '34410662' (subcomponent of...))
        bom_content (dict): Bom generated with bom_excel_to_dictionary
        modify_type (dict): {'type' : 'ADDED'}, {'type' : 'REMOVED'} or {'type' : 'CHANGED', value_changed:'Revision' 'new_value': '03', 'old_value': '02'}
        initial_dict (dict, optional): _description_. Defaults to {}.

    Returns:
        _type_: _description_
    """

    # Initialize a variable to hold the current level of the dictionary
    if initial_dict is None:
        initial_dict = {}

    current_level = initial_dict

    # Iterate through the keys to create the nested structure
    for i, key in enumerate(keys):
        if key not in current_level:
            # In this case, this specific key does not exist in the current level of the dictionary and should be created
            content = get_content(bom_content, keys[:i+1])  # get content of current item from full BOM
            if i == len(keys) - 1:
                # If we are at the last key, we can directly assign the modify_type
                mf = [modify_type]
            else:
                # If we are not at the last key, it means we are creating the subassembly that contains the modified item
                mf = [{'type': f'Item {modify_type['type']} inside'}]
            simplified_content = {'Description':  content['Description'], 'Revision': content['Revision'],
                                  'Quantity':  content['Quantity'], 'SupplyType':  content['SupplyType'], 'ModifyType': mf}

            current_level[key] = {'content': simplified_content}
        else:
            if i == len(keys) - 1:
                if modify_type not in current_level[key]['content']['ModifyType']:
                    current_level[key]['content']['ModifyType'].append(modify_type)
            else:
                new_item = {'type': f'Item {modify_type["type"]} inside'}
                if new_item not in current_level[key]['content']['ModifyType']:
                    current_level[key]['content']['ModifyType'].append(new_item)

        current_level = current_level[key]

    return initial_dict


def dict_to_table(d, maxDepth, level=1, result=None):
    """Transforms a dictionary BOM into a table like extracted from Oracle DB tables

    Args:
        d (dict): dictionnary to convert

    Returns:
        result: dict converted as list wiht level on 1rst row
    """
    if result is None:
        result = []
    for key, value in d.items():
        if key != 'content':
            result.append((level, *["" if i != level else level for i in range(1, maxDepth+1)],
                          key, *(d[key]['content'].values())))
            if isinstance(value, dict):
                dict_to_table(value, maxDepth, level=level + 1, result=result)
    return result


def get_content(bom: dict, keys: list):
    """get content of a dict bom from a string key

    Args:
        bom (dict): bom to analyse
        keys (list): list of key string ['34410718','34410662','34411697']
    """
    current_level = bom
    for i, key in enumerate(keys):
        if i == 0:
            current_level = current_level[key]
        else:
            current_level = current_level['content'][key]

    return current_level


def save_df_to_excel(result: DataFrame, max_depth: int, output_path: str, open_file=True):

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Convert DataFrame to rows and append to worksheet
    rows = dataframe_to_rows(result.drop("ModifyType", axis=1), index=False, header=True)
    # rows = dataframe_to_rows(result, index=False, header=True)
    header = next(rows)
    ws.append(header)

    for r in rows:
        # Convert numerical strings back to numbers and dict to strings
        r = [int(x) if isinstance(x, str) and x.isdigit() else (str(x) if isinstance(x, list) else x) for x in r]
        ws.append(r)

    # Left-align all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Set column widths to match the first file
    row_widths = [5] + [3] * max_depth + [31, 41, 8, 10, 19, 20]
    nb_columns = len(result.columns) - 1
    for i, width in enumerate(row_widths):
        ws.column_dimensions[chr(ord('A')+i)].width = width
        ws.column_dimensions[chr(ord('A')+i)].alignment = Alignment(horizontal='left')

    # Define table to written data
    tab = Table(displayName="Table1", ref=f'A1:{chr(ord('A')+nb_columns-1)}{result.shape[0]+1}')

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(tab)

    # Define the fill style with the desired background color
    fill_1 = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    fill_2 = PatternFill(start_color="DBDBDB", end_color="DBDBDB", fill_type="solid")
    fill_3 = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    fill_4 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_5 = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    fill = [fill_1, fill_2, fill_3, fill_4, fill_5]

    # Apply the fill style to the specified range
    for row in range(1, result.shape[0]+1):
        for col in range(1, nb_columns):
            ws.cell(row=row+1, column=col+1).fill = fill[col-1 if col <=
                                                         result.loc[row-1].at["Level"] else result.loc[row-1].at["Level"]-1]

    # Outline the specified range with the specified brush
    def outilne_range(min_row, min_col, max_row, max_col, brush):
        for row in range(min_row, max_row+1):
            ws.cell(row=row, column=min_col).border = Border(left=brush, bottom=(
                brush if row == max_row else None), top=(brush if row == min_row else None))
            ws.cell(row=row, column=max_col).border = Border(right=brush, bottom=(
                brush if row == max_row else None), top=(brush if row == min_row else None))
        for col in range(min_col+1, max_col):
            if max_row != min_row:
                ws.cell(row=min_row, column=col).border = Border(top=brush)
                ws.cell(row=max_row, column=col).border = Border(bottom=brush)
            else:
                ws.cell(row=min_row, column=col).border = Border(top=brush, bottom=brush)

    green_brush = Side(border_style="medium", color="2EB82E")
    red_brush = Side(border_style="medium", color="FF0000")

    bom1 = "bom1"
    bom2 = "bom2"

    # Create a persistent Tkinter root and font object at module level
    _tk_root = tk.Tk()
    _tk_root.withdraw()
    _default_font = font.Font(root=_tk_root, family="Calibri", size=11)

    def get_text_pixel_width(text, font_name="Calibri", font_size=11):
        # Use the persistent font object if default params, else create a new one
        if font_name == "Calibri" and font_size == 11:
            return _default_font.measure(text)
        else:
            tk_font = font.Font(root=_tk_root, family=font_name, size=font_size)
            return tk_font.measure(text)

    for index, row in result.iterrows():
        if {'type': 'ADDED'} in row['ModifyType']:
            # Green outline added items
            outilne_range(index+2, row['Level']+1, index+2, nb_columns, green_brush)
            comment_text = f'Item {result.iloc[index]["Item"]} was added in {bom2} (not present in {bom1})'
            length_splited_lines = [get_text_pixel_width(line) for line in comment_text.splitlines()]
            longest_line = max(length_splited_lines)
            max_line_length = 500
            nb_lines = sum([math.ceil((length)/max_line_length) for length in length_splited_lines])
            width = min(max_line_length, longest_line+10)  # 7px per char, up to 500px
            height = nb_lines * 18 + 8  # 20px per line, up to 500px
            ws[f'{chr(ord("A")+max_depth+1)}{index+2}'].comment = Comment(
                comment_text, "Automatically Generated", width=width, height=height)

        elif {'type': 'REMOVED'} in row['ModifyType']:
            # Red outline removed items
            outilne_range(index+2, row['Level']+1, index+2, nb_columns, red_brush)
            comment_text = f'Item {result.iloc[index]["Item"]} was removed in {bom2} (present in {bom1})'
            length_splited_lines = [get_text_pixel_width(line) for line in comment_text.splitlines()]
            longest_line = max(length_splited_lines)
            max_line_length = 500
            nb_lines = sum([math.ceil((length)/max_line_length) for length in length_splited_lines])
            width = min(max_line_length, longest_line+10)  # 7px per char, up to 500px
            height = nb_lines * 18 + 8  # 20px per line, up to 500px
            ws[f'{chr(ord("A")+max_depth+1)}{index+2}'].comment = Comment(
                comment_text, "Automatically Generated", width=width, height=height)
        else:
            for modif in row['ModifyType']:
                if modif['type'] == 'CHANGED':
                    comment_text = f"{modif['changed_value']} of item {result.iloc[index]['Item']} has changed :\n Old : {modif['old_value']}\n New: {modif['new_value']}"
                    # Dynamically set width and height, max width 500px
                    length_splited_lines = [get_text_pixel_width(line) for line in comment_text.splitlines()]
                    longest_line = max(length_splited_lines)
                    max_line_length = 500
                    nb_lines = sum([math.ceil((length)/max_line_length) for length in length_splited_lines])
                    width = min(max_line_length, longest_line+10)  # 7px per char, up to 500px
                    height = nb_lines * 18 + 8  # 20px per line, up to 500px
                    ws[f"{chr(ord('A')+result.columns.get_loc(modif['changed_value']))}{index+2}"].comment = Comment(
                        comment_text, "Automatically Generated", width=width, height=height)

        # else:
        #     for modif in row['ModifyType']:

    # SETTING CONFIDENTIALITY LABEL TO GENERAL
    props = CustomPropertyList()
    props.append(BoolProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_Enabled", value=True))
    props.append(DateTimeProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_SetDate", value="2024-12-23T12:50:17Z"))
    props.append(StringProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_Method", value="Privileged"))
    props.append(StringProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_Name", value="General v2"))
    props.append(StringProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_SiteId",
                 value="6e51e1ad-c54b-4b39-b598-0ffe9ae68fef"))
    props.append(StringProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_ActionId",
                 value="accb3366-0258-48d3-95a0-00a6be342c5b"))
    props.append(IntProperty(name="MSIP_Label_57443d00-af18-408c-9335-47b5de3ec9b9_ContentBits", value=2))

    # Assign the custom properties to the workbook
    wb.custom_doc_props = props

    # Save the workbook
    try:
        wb.save(output_path)
        if open_file:
            os.startfile(output_path)
    except PermissionError:
        raise PermissionError(
            f"Error: Cannot save file - {output_path} is already open. Please close it and try again.")


def compare_bom(path1: str, path2: str, output_path: str = None, open_result=True, simple_bom_mode=False):
    """
    Compare two Bill of Materials (BOM) Excel files.

    Args:
        path1 (str): The file path to the first BOM Excel file.
        path2 (str): The file path to the second BOM Excel file.

    Returns:
        None
    """
    if output_path is None:
        output_path = f"C:\\Users\\{os.getlogin()}\\Downloads\\compare{"_architecture_" if not simple_bom_mode else ""}result.xlsx"
    if not simple_bom_mode:
        bom1, m1 = bom_excel_to_dictionary(path1)
        bom2, m2 = bom_excel_to_dictionary(path2)
        max_depth = max(m1, m2)
    else:
        # Determine file type based on extension and use appropriate function
        if path1.lower().endswith('.xlsx'):
            bom1 = light_bom_excel_to_dictionary(path1)
        elif path1.lower().endswith('.txt'):
            bom1 = light_bom_oracle_to_dictionary(path1)

        if path2.lower().endswith('.xlsx'):
            bom2 = light_bom_excel_to_dictionary(path2)
        elif path2.lower().endswith('.txt'):
            bom2 = light_bom_oracle_to_dictionary(path2)
        max_depth = 1

    diff = DeepDiff(bom1, bom2, threshold_to_diff_deeper=0)

    item_added = [re.findall(r'\[\'(.*?)\'\]', element.replace('[\'content\']', "").replace("root", ""))
                  for element in diff.get('dictionary_item_added', [])]
    item_removed = [re.findall(r'\[\'(.*?)\'\]', element.replace('[\'content\']', "").replace("root", ""))
                    for element in diff.get('dictionary_item_removed', [])]
    item_changed = [(re.findall(r'\[\'(.*?)\'\]', element.replace('[\'content\']', "").replace("root", "")), diff['values_changed'][element])
                    for element in diff.get('values_changed', [])]
    type_changes = [(re.findall(r'\[\'(.*?)\'\]', element.replace('[\'content\']', "").replace("root", "")), diff['type_changes'][element])
                    for element in diff.get('type_changes', [])]
    # Combine type_changes and item_changed
    item_changed.extend((tc[0], {'new_value': tc[1]['new_value'], 'old_value': tc[1]['old_value']})
                        for tc in type_changes)
    print(item_changed)
    output = {}

    for item in item_added:
        output = append_to_dict(item, bom2, {'type': 'ADDED'}, output)

    for item in item_removed:
        output = append_to_dict(item, bom1, {'type': 'REMOVED'}, output)

    for item in item_changed:
        output = append_to_dict(item[0][:-1], bom2, {'type': 'CHANGED',
                                'changed_value': item[0][-1], **item[1]}, output)

    table_output = dict_to_table(output, max_depth)

    columns = ['Level', *[str(i) for i in range(1, max_depth+1)], 'Item', 'Description', 'Revision',
               'Quantity', 'SupplyType', 'ModifyType']
    # Transform ModifyType into readable string

    def transform_modify_type(modify_type_list):
        messages = []
        for item in modify_type_list:
            if item['type'] == 'REMOVED':
                messages.append('Removed')
            elif item['type'] == 'ADDED':
                messages.append('Added')
            elif item['type'] == 'CHANGED':
                messages.append(f"Changed {item['changed_value']}")
            elif 'Item' in item['type']:
                messages.append(item['type'])
        return ', '.join(messages)

    output_df = DataFrame(table_output, columns=columns)

    # Apply transformation to ModifyType column
    output_df['Changed'] = output_df['ModifyType'].apply(transform_modify_type)

    if item_added or item_changed or item_removed:
        save_df_to_excel(output_df, max_depth, output_path, open_file=open_result)
        return True
    else:
        return False

# Function to check if a file exists


def check_file(file_path):
    if os.path.isfile(file_path):
        print(f"File exists: {file_path}")
    else:
        print(f"File does not exist: {file_path}")


##################################################### WORKING CODE #################################################

##################################################### WORKING CODE #################################################

# file_1 = "C:/Users/SESA787052/Downloads/BOM QBOT21000 _rev02 1.xlsx"
# # file_2 = "C:/Users/SESA787052/Downloads/fnd_gfm_204718568.txt"
# file_2 = "./fnd_gfm_204718568.txt"

# try:
#     compare_bom(file_1, file_2, simple_bom_mode=True)
# except PermissionError as e:
#     print(str(e))
